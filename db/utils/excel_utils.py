from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


def create_payslip_excel(employee_name, cedula, fecha_pago, detalle):
    """Generate a Spanish payslip (recibo de pago) in Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Recibo de Pago"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # ─ Header ──────────────────────────────────────────────────────────────────
    ws['A1'] = "RECIBO DE PAGO"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True)
    ws.merge_cells('A1:D1')

    row = 3
    ws[f'A{row}'] = "Empleado:"
    ws[f'B{row}'] = employee_name
    ws[f'B{row}'].font = Font(bold=True)
    ws.merge_cells(f'B{row}:D{row}')

    row += 1
    ws[f'A{row}'] = "Cédula:"
    ws[f'B{row}'] = cedula

    row += 1
    ws[f'A{row}'] = "Fecha de Pago:"
    ws[f'B{row}'] = datetime.strptime(fecha_pago, "%Y-%m-%d").strftime("%d/%m/%Y")

    row += 1
    ws[f'A{row}'] = "Tasa USD:"
    ws[f'B{row}'] = detalle['tasa_dolar']
    ws[f'B{row}'].number_format = '0.00'

    # ─ Ingresos (LEFT) ─────────────────────────────────────────────────────────
    row += 2
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    ws[f'A{row}'] = "INGRESOS"
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:B{row}')

    ingresos_start = row + 1
    row += 1

    ws[f'A{row}'] = "Salario Base"
    ws[f'B{row}'] = detalle['salario_base']
    ws[f'B{row}'].number_format = '#,##0.00'
    row += 1

    for bono in detalle['bonos']:
        ws[f'A{row}'] = bono['nombre']
        ws[f'B{row}'] = bono['monto_aplicado']
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1

    # Cestaticket (if present)
    if detalle.get('cestaticket_aplicado') and float(detalle['cestaticket_aplicado']) > 0:
        ws[f'A{row}'] = "Cesta Ticket"
        ws[f'B{row}'] = float(detalle['cestaticket_aplicado'])
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1

    ingresos_total_row = row
    ws[f'A{row}'] = "TOTAL INGRESOS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"=SUM(B{ingresos_start}:B{row - 1})"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # ─ Deducciones (RIGHT) ─────────────────────────────────────────────────────
    drow = ingresos_start - 2
    ws[f'D{drow}'] = "DEDUCCIONES"
    ws[f'D{drow}'].fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    ws[f'D{drow}'].font = Font(color="FFFFFF", bold=True, size=11)
    ws.merge_cells(f'D{drow}:E{drow}')

    drow += 1
    deducciones_start = drow

    for ded in detalle['deducciones']:
        ws[f'D{drow}'] = ded['nombre']
        ws[f'E{drow}'] = ded['monto_aplicado']
        ws[f'E{drow}'].number_format = '#,##0.00'
        drow += 1

    deducciones_total_row = drow
    ws[f'D{drow}'] = "TOTAL DEDUCCIONES"
    ws[f'D{drow}'].font = Font(bold=True)
    ws[f'E{drow}'] = f"=SUM(E{deducciones_start}:E{drow - 1})"
    ws[f'E{drow}'].font = Font(bold=True)
    ws[f'E{drow}'].number_format = '#,##0.00'
    ws[f'E{drow}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # ─ Net (Bottom) ────────────────────────────────────────────────────────────
    summary_row = max(ingresos_total_row, deducciones_total_row) + 2

    ws[f'A{summary_row}'] = "SALARIO NETO"
    ws[f'A{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'] = f"=B{ingresos_total_row}-E{deducciones_total_row}"
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].number_format = '#,##0.00'
    ws[f'B{summary_row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    summary_row += 1
    ws[f'A{summary_row}'] = "SALARIO NETO (USD)"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'] = f"=B{summary_row-1}/B5"
    ws[f'B{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].number_format = '$#,##0.00'
    ws[f'B{summary_row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    # ─ Formatting ──────────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_payroll_report_excel(nomina_id, detalle_empleados, fecha_pago, tasa_dolar):
    """Generate a full payroll report in Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Nómina"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    # ─ Report Title ────────────────────────────────────────────────────────────
    ws['A1'] = f"NÓMINA - {fecha_pago}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')

    ws['A2'] = f"Tasa USD: {tasa_dolar:.2f}"
    ws.merge_cells('A2:C2')

    # ─ Table Headers ───────────────────────────────────────────────────────────
    row = 4
    headers = ["Cédula", "Nombre", "Salario Base", "Bonos", "Cesta Ticket", "Deducciones", "Total Ingresos", "Salario Final (Bs)", "Salario Final (USD)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ─ Data Rows ───────────────────────────────────────────────────────────────
    row = 5
    for emp in detalle_empleados:
        ws.cell(row=row, column=1).value = emp['empleado_cedula']
        ws.cell(row=row, column=2).value = emp['nombre']
        ws.cell(row=row, column=3).value = float(emp['salario_base'])
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4).value = float(emp['total_ingresos']) - float(emp['salario_base'])
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5).value = float(emp.get('cestaticket_aplicado', 0))
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=6).value = float(emp['total_deducciones'])
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=7).value = float(emp['total_ingresos'])
        ws.cell(row=row, column=7).number_format = '#,##0.00'
        ws.cell(row=row, column=8).value = float(emp['salario_final_bs'])
        ws.cell(row=row, column=8).number_format = '#,##0.00'
        ws.cell(row=row, column=9).value = float(emp['salario_final_bs']) / tasa_dolar
        ws.cell(row=row, column=9).number_format = '$#,##0.00'
        row += 1

    # ─ Totals ──────────────────────────────────────────────────────────────────
    totals_row = row
    ws.cell(row=row, column=1).value = "TOTALES"
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col in range(3, 9):
        cell = ws.cell(row=row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{row-1})"
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        if col in [3, 4, 5, 6, 7]:
            cell.number_format = '#,##0.00'
        else:
            cell.number_format = '$#,##0.00'

    # ─ Column widths ───────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_liquidacion_excel(liquidacion: dict) -> bytes:
    """Genera el recibo de liquidación en Excel con desglose completo Art. 142 LOTTT."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Recibo de Liquidación"

    blue_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    orange_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    green_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    bold = Font(bold=True)
    money_fmt = '#,##0.00'

    # ── Título ────────────────────────────────────────────────────────────────────
    ws['A1'] = "RECIBO DE LIQUIDACIÓN"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')

    row = 3
    datos = [
        ("Empleado:", liquidacion.get("empleado_nombre", liquidacion.get("empleado_cedula"))),
        ("Cédula:", liquidacion.get("empleado_cedula")),
        ("Salario Base (Bs):", liquidacion.get("empleado_salario_base")),
        ("Fecha de Egreso:", liquidacion.get("fecha_egreso")),
        ("Causa de Egreso:", liquidacion.get("causa_egreso")),
        ("Años de Servicio:", f"{liquidacion.get('anios_totales', 0):.2f}"),
        ("Tasa USD (Bs):", liquidacion.get("tasa_dolar")),
        ("Salario Integral Diario (Bs):", liquidacion.get("salario_integral_dia")),
    ]
    for label, valor in datos:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = bold
        ws[f'B{row}'] = valor
        ws.merge_cells(f'B{row}:F{row}')
        row += 1

    row += 1

    def section_header(title: str, fill):
        nonlocal row
        ws[f'A{row}'] = title
        ws[f'A{row}'].fill = fill
        ws[f'A{row}'].font = white_font
        ws.merge_cells(f'A{row}:F{row}')

    def col_headers():
        nonlocal row
        row += 1
        for col, h in enumerate(["Concepto", "Días", "SID (Bs)", "Monto (Bs)", "Monto (USD)", "Observación"], 1):
            cell = ws.cell(row=row, column=col)
            cell.value = h
            cell.font = bold
            cell.fill = gray_fill
        row += 1

    def data_row(concepto, dias, sid, monto_bs, monto_usd, obs=""):
        nonlocal row
        ws.cell(row=row, column=1).value = concepto
        ws.cell(row=row, column=2).value = round(dias, 2) if dias is not None else "—"
        ws.cell(row=row, column=3).value = round(sid, 4) if sid else "—"
        ws.cell(row=row, column=4).value = monto_bs
        ws.cell(row=row, column=4).number_format = money_fmt
        ws.cell(row=row, column=5).value = monto_usd
        ws.cell(row=row, column=5).number_format = money_fmt
        ws.cell(row=row, column=6).value = obs or ""
        row += 1

    def subtotal_row(label, monto_bs, monto_usd):
        nonlocal row
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = bold
        ws.cell(row=row, column=4).value = monto_bs
        ws.cell(row=row, column=4).font = bold
        ws.cell(row=row, column=4).number_format = money_fmt
        ws.cell(row=row, column=4).fill = gray_fill
        ws.cell(row=row, column=5).value = monto_usd
        ws.cell(row=row, column=5).font = bold
        ws.cell(row=row, column=5).number_format = money_fmt
        ws.cell(row=row, column=5).fill = gray_fill
        row += 1

    sid = liquidacion.get("salario_integral_dia") or 0
    tasa = liquidacion.get("tasa_dolar") or 1

    desglose = {d["concepto"]: d for d in liquidacion.get("desglose", [])}

    def d(concepto):
        return desglose.get(concepto, {})

    # ── Prestaciones Sociales Art. 142 ───────────────────────────────────────────
    section_header("PRESTACIONES SOCIALES (Art. 142 LOTTT)", blue_fill)
    col_headers()

    esc_a = liquidacion.get("escenario_a_bs") or 0
    esc_b = liquidacion.get("escenario_b_bs") or 0
    esc = liquidacion.get("escenario_aplicado", "?")
    prestaciones_bs = liquidacion.get("prestaciones_bs") or 0

    data_row("  Escenario A — Depósitos trimestrales + adicionales",
             None, sid, esc_a, round(esc_a / tasa, 2),
             f"{'✓ APLICADO' if esc == 'A' else ''}")
    data_row("  Escenario B — SID × 30 días × años de servicio",
             None, sid, esc_b, round(esc_b / tasa, 2),
             f"{'✓ APLICADO' if esc == 'B' else ''}")
    subtotal_row(f"Prestaciones Sociales (Escenario {esc})", prestaciones_bs, round(prestaciones_bs / tasa, 2))

    row += 1

    # ── Conceptos Fraccionados ────────────────────────────────────────────────────
    section_header("CONCEPTOS FRACCIONADOS (Ingresos)", blue_fill)
    col_headers()

    conceptos = [
        ("Vacaciones Fraccionadas", "vacaciones_fracc_bs", "vacaciones_fraccionadas"),
        ("Bono Vacacional Fraccionado", "bono_vac_fracc_bs", "bono_vac_fraccionado"),
        ("Utilidades Fraccionadas", "utilidades_fracc_bs", "utilidades_fraccionadas"),
        ("Salarios Pendientes", "salarios_pendientes_bs", "salarios_pendientes"),
        ("Intereses de Prestaciones", "intereses_bs", "intereses"),
    ]

    subtotal_fracc = 0
    for label, field, concepto_key in conceptos:
        monto = liquidacion.get(field) or 0
        linea = d(concepto_key)
        dias = linea.get("cantidad_dias")
        obs = linea.get("observacion", "")
        data_row(label, dias, sid if dias else None, monto, round(monto / tasa, 2), obs)
        subtotal_fracc += monto

    subtotal_row("Subtotal Conceptos Fraccionados", round(subtotal_fracc, 2), round(subtotal_fracc / tasa, 2))

    row += 1

    # ── Deducciones ───────────────────────────────────────────────────────────────
    section_header("DEDUCCIONES", orange_fill)
    col_headers()

    deudor = liquidacion.get("saldo_deudor_bs") or 0
    data_row("Saldo Deudor Préstamos y Anticipos", None, None, deudor, round(deudor / tasa, 2),
             d("saldo_deudor").get("observacion", ""))
    subtotal_row("Total Deducciones", deudor, round(deudor / tasa, 2))

    row += 1

    # ── Resumen Final ─────────────────────────────────────────────────────────────
    section_header("RESUMEN FINAL", green_fill)
    row += 1

    monto_total_bs = liquidacion.get("monto_total_bs") or 0
    monto_total_usd = liquidacion.get("monto_total_usd") or 0
    monto_neto_bs = liquidacion.get("monto_neto_bs") or 0
    monto_neto_usd = liquidacion.get("monto_neto_usd") or 0

    resumen = [
        ("Total Bruto (Bs)", monto_total_bs, money_fmt),
        ("Total Bruto (USD)", monto_total_usd, money_fmt),
        ("(-) Deducciones (Bs)", deudor, money_fmt),
        ("NETO A PAGAR (Bs)", monto_neto_bs, money_fmt),
        ("NETO A PAGAR (USD)", monto_neto_usd, money_fmt),
    ]
    for label, valor, fmt in resumen:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = bold
        ws[f'D{row}'] = valor
        ws[f'D{row}'].font = Font(bold=True, size=12 if "NETO" in label else 11)
        ws[f'D{row}'].number_format = fmt
        if "NETO A PAGAR" in label:
            ws[f'D{row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        row += 1

    # ── Ancho de columnas ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 45

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
